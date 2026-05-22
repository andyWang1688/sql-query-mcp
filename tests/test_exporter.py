from __future__ import annotations

import csv
import json
import tempfile
import unittest
import uuid
from contextlib import contextmanager
from datetime import datetime, timezone
from pathlib import Path
from typing import cast

from openpyxl import load_workbook

from sql_query_mcp.audit import AuditLogger
from sql_query_mcp.config import ConnectionConfig, ServerSettings
from sql_query_mcp.errors import QueryExecutionError
from sql_query_mcp.exporter import QueryExporter


class _CursorStub:
    def __init__(self, batches, description=("id", "name")) -> None:
        self._batches = list(batches)
        self.description = description
        self.executed = []
        self.fetchmany_sizes = []

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc, tb) -> None:
        return None

    def execute(self, query) -> None:
        self.executed.append(query)

    def fetchmany(self, size: int):
        self.fetchmany_sizes.append(size)
        if self._batches:
            return self._batches.pop(0)
        return []


class _ConnectionStub:
    def __init__(self, cursor: _CursorStub) -> None:
        self._cursor = cursor

    def cursor(self):
        return self._cursor


class _AdapterStub:
    def __init__(self) -> None:
        self.set_statement_timeout_calls = []

    def set_statement_timeout(self, conn: object, timeout_ms: int) -> None:
        self.set_statement_timeout_calls.append(timeout_ms)

    def column_names(self, description):
        return list(description or [])


class _StreamingAdapterStub(_AdapterStub):
    def __init__(self, export_cursor) -> None:
        super().__init__()
        self._export_cursor = export_cursor
        self.export_cursor_calls = 0

    def export_cursor(self, conn: object):
        self.export_cursor_calls += 1
        return self._export_cursor


class _RegistryStub:
    def __init__(self, config: ConnectionConfig, adapter: object, conn: object) -> None:
        self._config = config
        self._adapter = adapter
        self._conn = conn
        self.connection_calls = 0

    def get_connection_config(self, connection_id: str) -> ConnectionConfig:
        if connection_id != self._config.connection_id:
            raise AssertionError(connection_id)
        return self._config

    @contextmanager
    def connection_from_config(self, config: ConnectionConfig):
        if config != self._config:
            raise AssertionError(config)
        self.connection_calls += 1
        yield self._conn, self._adapter


def _mysql_config() -> ConnectionConfig:
    return ConnectionConfig(
        connection_id="crm_mysql_prod_main_ro",
        engine="mysql",
        label="CRM MySQL",
        env="prod",
        tenant="main",
        role="ro",
        dsn_env="MYSQL_CONN",
        enabled=True,
        default_database="crm",
    )


def _postgres_config() -> ConnectionConfig:
    return ConnectionConfig(
        connection_id="crm_pg_prod_main_ro",
        engine="postgres",
        label="CRM PostgreSQL",
        env="prod",
        tenant="main",
        role="ro",
        dsn_env="PG_CONN",
        enabled=True,
        default_schema="public",
    )


def _hive_config() -> ConnectionConfig:
    return ConnectionConfig(
        connection_id="warehouse_hive_prod_main_ro",
        engine="hive",
        label="Warehouse Hive",
        env="prod",
        tenant="main",
        role="ro",
        dsn_env="HIVE_CONN",
        enabled=True,
        default_database="default",
    )


def _build_exporter(
    config: ConnectionConfig,
    cursor: _CursorStub,
    temp_dir: str,
    adapter: _AdapterStub | None = None,
):
    adapter = adapter or _AdapterStub()
    registry = _RegistryStub(config, adapter, _ConnectionStub(cursor))
    log_path = Path(temp_dir) / "audit.jsonl"
    exporter = QueryExporter(
        registry=registry,
        settings=ServerSettings(
            default_limit=200,
            max_limit=1000,
            statement_timeout_ms=2500,
            audit_log_path=log_path,
        ),
        audit_logger=AuditLogger(log_path),
    )
    return exporter, registry, adapter, log_path


def _read_audit_records(path: Path):
    return [json.loads(line) for line in path.read_text(encoding="utf-8").splitlines()]


class QueryExporterTestCase(unittest.TestCase):
    def test_export_csv_writes_multiple_fetchmany_batches(self) -> None:
        cursor = _CursorStub(
            batches=[
                [(1, "Alice")],
                [(2, "Bob")],
                [],
            ]
        )
        with tempfile.TemporaryDirectory() as temp_dir:
            exporter, _, adapter, log_path = _build_exporter(
                _mysql_config(), cursor, temp_dir
            )

            result = exporter.export_query_file(
                "crm_mysql_prod_main_ro",
                "SELECT id, name FROM users",
                temp_dir,
                limit=2,
                file_name="users",
            )
            file_path = cast(str, result["file_path"])
            with Path(file_path).open("r", encoding="utf-8", newline="") as handle:
                rows = list(csv.reader(handle))
            records = _read_audit_records(log_path)

        self.assertEqual("csv", result["format"])
        self.assertEqual(2, result["row_count"])
        self.assertEqual(2, result["applied_limit"])
        self.assertEqual(["id", "name"], rows[0])
        self.assertEqual([["1", "Alice"], ["2", "Bob"]], rows[1:])
        self.assertIn("LIMIT 2", cursor.executed[0])
        self.assertEqual([2500], adapter.set_statement_timeout_calls)
        self.assertGreaterEqual(len(cursor.fetchmany_sizes), 2)
        self.assertEqual("export_query_file", records[0]["tool"])
        self.assertEqual(2, records[0]["row_count"])

    def test_export_xlsx_writes_rows(self) -> None:
        cursor = _CursorStub(batches=[[(1, "Alice")], []])
        with tempfile.TemporaryDirectory() as temp_dir:
            exporter, _, _, _ = _build_exporter(_postgres_config(), cursor, temp_dir)

            result = exporter.export_query_file(
                "crm_pg_prod_main_ro",
                "SELECT id, name FROM users",
                temp_dir,
                format="xlsx",
                file_name="users",
            )
            workbook = load_workbook(cast(str, result["file_path"]), read_only=True)
            worksheet = workbook.active
            if worksheet is None:
                raise AssertionError("workbook has no active worksheet")
            rows = list(worksheet.iter_rows(values_only=True))
            workbook.close()

        self.assertEqual("xlsx", result["format"])
        self.assertEqual([("id", "name"), (1, "Alice")], rows)

    def test_export_all_does_not_wrap_query_with_limit(self) -> None:
        cursor = _CursorStub(batches=[[(1, "Alice")], []])
        with tempfile.TemporaryDirectory() as temp_dir:
            exporter, _, _, _ = _build_exporter(_mysql_config(), cursor, temp_dir)
            result = exporter.export_query_file(
                "crm_mysql_prod_main_ro",
                "SELECT id, name FROM users",
                temp_dir,
                export_all=True,
                file_name="users",
            )

        self.assertEqual("SELECT id, name FROM users", cursor.executed[0])
        self.assertIsNone(result["applied_limit"])

    def test_empty_result_creates_header_only_csv(self) -> None:
        cursor = _CursorStub(batches=[[]])
        with tempfile.TemporaryDirectory() as temp_dir:
            exporter, _, _, _ = _build_exporter(_mysql_config(), cursor, temp_dir)
            result = exporter.export_query_file(
                "crm_mysql_prod_main_ro",
                "SELECT id, name FROM users",
                temp_dir,
                file_name="empty",
            )
            file_path = cast(str, result["file_path"])
            with Path(file_path).open("r", encoding="utf-8", newline="") as handle:
                rows = list(csv.reader(handle))

        self.assertEqual(0, result["row_count"])
        self.assertEqual([["id", "name"]], rows)

    def test_existing_file_gets_numbered_name_when_not_overwriting(self) -> None:
        cursor = _CursorStub(batches=[[]])
        with tempfile.TemporaryDirectory() as temp_dir:
            existing = Path(temp_dir) / "users.csv"
            existing.write_text("old", encoding="utf-8")
            exporter, _, _, _ = _build_exporter(_mysql_config(), cursor, temp_dir)
            result = exporter.export_query_file(
                "crm_mysql_prod_main_ro",
                "SELECT id, name FROM users",
                temp_dir,
                file_name="users",
            )

        self.assertTrue(cast(str, result["file_path"]).endswith("users (1).csv"))

    def test_overwrite_replaces_existing_file(self) -> None:
        cursor = _CursorStub(batches=[[]])
        with tempfile.TemporaryDirectory() as temp_dir:
            existing = Path(temp_dir) / "users.csv"
            existing.write_text("old", encoding="utf-8")
            exporter, _, _, _ = _build_exporter(_mysql_config(), cursor, temp_dir)
            result = exporter.export_query_file(
                "crm_mysql_prod_main_ro",
                "SELECT id, name FROM users",
                str(existing),
                overwrite=True,
            )
            content = existing.read_text(encoding="utf-8")

        self.assertEqual(str(existing), result["file_path"])
        self.assertNotEqual("old", content)

    def test_invalid_sql_rejected_before_connecting(self) -> None:
        cursor = _CursorStub(batches=[])
        with tempfile.TemporaryDirectory() as temp_dir:
            exporter, registry, _, _ = _build_exporter(_mysql_config(), cursor, temp_dir)
            with self.assertRaises(QueryExecutionError):
                exporter.export_query_file(
                    "crm_mysql_prod_main_ro",
                    "DELETE FROM users",
                    temp_dir,
                )

        self.assertEqual(0, registry.connection_calls)

    def test_hive_rejected_before_connecting(self) -> None:
        cursor = _CursorStub(batches=[])
        with tempfile.TemporaryDirectory() as temp_dir:
            exporter, registry, _, _ = _build_exporter(_hive_config(), cursor, temp_dir)
            with self.assertRaises(QueryExecutionError):
                exporter.export_query_file(
                    "warehouse_hive_prod_main_ro",
                    "SELECT id, name FROM users",
                    temp_dir,
                )

        self.assertEqual(0, registry.connection_calls)

    def test_export_uses_adapter_export_cursor_when_available(self) -> None:
        default_cursor = _CursorStub(batches=[])
        export_cursor = _CursorStub(batches=[[(1, "Alice")], []])
        adapter = _StreamingAdapterStub(export_cursor=export_cursor)
        with tempfile.TemporaryDirectory() as temp_dir:
            exporter, _, _, _ = _build_exporter(
                _mysql_config(), default_cursor, temp_dir, adapter=adapter
            )

            result = exporter.export_query_file(
                "crm_mysql_prod_main_ro",
                "SELECT id, name FROM users",
                temp_dir,
                file_name="users",
            )

        self.assertEqual(1, adapter.export_cursor_calls)
        self.assertEqual(1, result["row_count"])
        self.assertEqual([], default_cursor.executed)
        self.assertEqual(1, len(export_cursor.executed))

    def test_export_xlsx_converts_uuid_values_to_text(self) -> None:
        row_id = uuid.uuid4()
        cursor = _CursorStub(batches=[[(row_id, "Alice")], []])
        with tempfile.TemporaryDirectory() as temp_dir:
            exporter, _, _, _ = _build_exporter(_postgres_config(), cursor, temp_dir)

            result = exporter.export_query_file(
                "crm_pg_prod_main_ro",
                "SELECT id, name FROM users",
                temp_dir,
                format="xlsx",
                file_name="users",
            )
            workbook = load_workbook(cast(str, result["file_path"]), read_only=True)
            worksheet = workbook.active
            if worksheet is None:
                raise AssertionError("workbook has no active worksheet")
            rows = list(worksheet.iter_rows(values_only=True))
            workbook.close()

        self.assertEqual(str(row_id), rows[1][0])

    def test_export_xlsx_removes_datetime_timezone(self) -> None:
        value = datetime(2026, 1, 1, 8, 30, tzinfo=timezone.utc)
        cursor = _CursorStub(batches=[[(value, "Alice")], []], description=("update_time", "name"))
        with tempfile.TemporaryDirectory() as temp_dir:
            exporter, _, _, _ = _build_exporter(_postgres_config(), cursor, temp_dir)

            result = exporter.export_query_file(
                "crm_pg_prod_main_ro",
                "SELECT update_time, name FROM users",
                temp_dir,
                format="xlsx",
                file_name="users",
            )
            workbook = load_workbook(cast(str, result["file_path"]), read_only=True)
            worksheet = workbook.active
            if worksheet is None:
                raise AssertionError("workbook has no active worksheet")
            rows = list(worksheet.iter_rows(values_only=True))
            workbook.close()

        self.assertEqual(value.replace(tzinfo=None), rows[1][0])

    def test_export_xlsx_error_includes_column_and_type(self) -> None:
        cursor = _CursorStub(batches=[[(object(), "Alice")], []])
        with tempfile.TemporaryDirectory() as temp_dir:
            exporter, _, _, _ = _build_exporter(_postgres_config(), cursor, temp_dir)

            with self.assertRaises(QueryExecutionError) as caught:
                exporter.export_query_file(
                    "crm_pg_prod_main_ro",
                    "SELECT id, name FROM users",
                    temp_dir,
                    format="xlsx",
                    file_name="users",
                )

        self.assertIn("id", str(caught.exception))
        self.assertIn("object", str(caught.exception))


if __name__ == "__main__":
    unittest.main()
