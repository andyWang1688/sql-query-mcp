"""Controlled local file imports into existing tables."""

from __future__ import annotations

import csv
import time
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - runtime dependency
    load_workbook = None

from .audit import AuditLogger
from .config import ServerSettings
from .errors import QueryExecutionError, sanitize_error_message
from .namespace import NamespaceSelection, resolve_namespace


class TableFileImporter:
    """Import CSV and XLSX files through a constrained table-only path."""

    def __init__(
        self,
        registry: Any,
        settings: ServerSettings,
        audit_logger: AuditLogger,
    ):
        self._registry = registry
        self._settings = settings
        self._audit = audit_logger

    def import_table_file(
        self,
        connection_id: str,
        table_name: str,
        file_path: str,
        schema: Optional[str] = None,
        database: Optional[str] = None,
        sheet_name: Optional[str] = None,
    ) -> Dict[str, object]:
        started = time.perf_counter()
        config = None
        namespace = None
        file_extension = Path(file_path).suffix.lower()
        selected_sheet_name = None
        inserted_row_count = 0

        try:
            config = self._registry.get_connection_config(connection_id)
            namespace = resolve_namespace(config, schema=schema, database=database)
            headers, rows, selected_sheet_name = _read_file(Path(file_path), sheet_name)
            if not rows:
                raise QueryExecutionError("文件没有可导入的数据行。")

            with self._registry.connection_from_config(config) as (conn, adapter):
                _apply_statement_timeout(adapter, conn, self._settings.statement_timeout_ms)
                description = adapter.describe_table(conn, namespace.value, table_name)
                if not description:
                    raise QueryExecutionError(
                        f"未找到表 {namespace.value}.{table_name}，或当前用户没有访问权限"
                    )
                table_columns = [item["column_name"] for item in description["columns"]]
                _validate_headers(headers, table_columns)
                query = adapter.build_insert_query(namespace.value, table_name, headers)
                _execute_insert(conn, config.engine, query, rows)

            inserted_row_count = len(rows)
            duration_ms = _elapsed_ms(started)
            self._audit.log(
                tool="import_table_file",
                connection_id=connection_id,
                success=True,
                duration_ms=duration_ms,
                row_count=inserted_row_count,
                extra=_build_audit_extra(
                    config,
                    namespace,
                    table_name,
                    file_extension,
                    selected_sheet_name,
                ),
            )
            return {
                "connection_id": connection_id,
                "engine": config.engine,
                namespace.field_name: namespace.value,
                "table_name": table_name,
                "inserted_row_count": inserted_row_count,
                "duration_ms": duration_ms,
                "file_extension": file_extension,
                "sheet_name": selected_sheet_name,
            }
        except Exception as exc:
            duration_ms = _elapsed_ms(started)
            sanitized = sanitize_error_message(str(exc))
            self._audit.log(
                tool="import_table_file",
                connection_id=connection_id,
                success=False,
                duration_ms=duration_ms,
                row_count=inserted_row_count,
                error=sanitized,
                extra=_build_audit_extra(
                    config,
                    namespace,
                    table_name,
                    file_extension,
                    selected_sheet_name,
                ),
            )
            raise QueryExecutionError(sanitized) from exc


def _read_file(path: Path, sheet_name: Optional[str]) -> Tuple[List[str], List[Tuple[object, ...]], Optional[str]]:
    extension = path.suffix.lower()
    if extension == ".csv":
        if sheet_name:
            raise QueryExecutionError("CSV 文件不支持 sheet_name 参数。")
        return _read_csv(path)
    if extension == ".xlsx":
        return _read_xlsx(path, sheet_name)
    raise QueryExecutionError("仅支持 .csv 和 .xlsx 文件导入。")


def _read_csv(path: Path) -> Tuple[List[str], List[Tuple[object, ...]], Optional[str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.reader(handle)
        try:
            headers = next(reader)
        except StopIteration as exc:
            raise QueryExecutionError("文件表头不能为空。") from exc
        rows = [_normalize_row(row, len(headers)) for row in reader]
    return headers, rows, None


def _read_xlsx(path: Path, sheet_name: Optional[str]) -> Tuple[List[str], List[Tuple[object, ...]], Optional[str]]:
    if load_workbook is None:
        raise QueryExecutionError("缺少 openpyxl 依赖，请先安装项目依赖。")
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name:
            if sheet_name not in workbook.sheetnames:
                raise QueryExecutionError(f"XLSX 文件中不存在 sheet: {sheet_name}")
            worksheet = workbook[sheet_name]
        else:
            worksheet = workbook.worksheets[0]
        rows_iter = worksheet.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration as exc:
            raise QueryExecutionError("文件表头不能为空。") from exc
        headers = ["" if value is None else str(value) for value in header_row]
        rows = [_normalize_row(list(row), len(headers)) for row in rows_iter]
        return headers, rows, worksheet.title
    finally:
        workbook.close()


def _normalize_row(row: Sequence[object], expected_length: int) -> Tuple[object, ...]:
    if len(row) != expected_length:
        raise QueryExecutionError("数据行字段数量必须和表头数量一致。")
    return tuple(None if value == "" else value for value in row)


def _validate_headers(headers: Sequence[str], table_columns: Sequence[str]) -> None:
    if not headers:
        raise QueryExecutionError("文件表头不能为空。")
    empty_headers = [index + 1 for index, header in enumerate(headers) if not header]
    if empty_headers:
        raise QueryExecutionError(f"文件表头存在空字段，位置: {empty_headers}")
    duplicates = sorted({header for header in headers if headers.count(header) > 1})
    if duplicates:
        raise QueryExecutionError(f"文件表头存在重复字段: {duplicates}")
    unknown = sorted(set(headers) - set(table_columns))
    if unknown:
        raise QueryExecutionError(f"文件表头包含目标表不存在的字段: {unknown}")


def _execute_insert(conn: Any, engine: str, query: object, rows: List[Tuple[object, ...]]) -> None:
    if engine == "postgres" and hasattr(conn, "transaction"):
        with conn.transaction():
            with conn.cursor() as cur:
                cur.executemany(query, rows)
        return

    if engine == "hive" and not hasattr(conn, "begin"):
        with conn.cursor() as cur:
            cur.executemany(query, rows)
        return

    conn.begin()
    try:
        with conn.cursor() as cur:
            cur.executemany(query, rows)
        conn.commit()
    except Exception:
        conn.rollback()
        raise


def _elapsed_ms(started: float) -> int:
    return int((time.perf_counter() - started) * 1000)


def _apply_statement_timeout(adapter: Any, conn: Any, timeout_ms: Optional[int]) -> None:
    if timeout_ms is not None:
        getattr(adapter, "set_statement_timeout")(conn, timeout_ms)


def _build_audit_extra(
    config: Any,
    namespace: Optional[NamespaceSelection],
    table_name: str,
    file_extension: str,
    sheet_name: Optional[str],
) -> Dict[str, object]:
    extra: Dict[str, object] = {
        "table_name": table_name,
        "file_extension": file_extension,
        "sheet_name": sheet_name,
    }
    if config is not None:
        extra["engine"] = config.engine
    if namespace is not None:
        extra[namespace.field_name] = namespace.value
    return extra
